import openpyxl

path = "услуги-тарифы-2026.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
for ws in wb.worksheets:
    print(f"=== Лист: {ws.title}  ({ws.max_row}x{ws.max_column}) ===")
    for row in ws.iter_rows(values_only=True):
        print(row)
    print()
